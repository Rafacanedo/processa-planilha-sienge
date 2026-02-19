"""
Pipeline: Transform planilha_leonardo.xlsx → planilha_final.xlsx

Normalizes item hierarchy to always have 4 levels (XXX.XXX.XXX.XXX).
Data items must be at level 4; missing intermediate levels are auto-created.
"""

import sys
import re
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl


# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------

@dataclass
class InputItem:
    """A parsed row from the input spreadsheet."""
    raw_item: str          # e.g. "01.02.01" or "4"
    description: str
    code: str | None       # CÓDIGO  — None for header rows
    unit: str | None       # UNID.
    price: float | None    # PREÇO UNITÁRIO
    quantity: float | None # QUANTIDADE (col S)
    is_data: bool = False  # True when item has code + unit (leaf item)

    @property
    def parts(self) -> list[str]:
        """Split item into its dotted parts, zero-padded to 3 digits."""
        raw = str(self.raw_item)
        # Level-1 items may be plain integers (e.g. 1, 2, 3)
        segments = raw.split(".")
        return [seg.zfill(3) for seg in segments]

    @property
    def level(self) -> int:
        return len(self.parts)

    @property
    def padded_item(self) -> str:
        return ".".join(self.parts)


@dataclass
class OutputRow:
    """A row to write in the output spreadsheet."""
    item: str               # e.g. "001.002.001.001"
    code: str | None = None
    description: str = ""
    unit: str | None = None
    quantity: float | None = None
    price: float | None = None


# ---------------------------------------------------------------------------
# Unit normalisation
# ---------------------------------------------------------------------------

UNIT_MAP = {
    "M2":  "m2",
    "M3":  "m3",
    "M":   "m",
    "KG":  "kg",
    "UND": "un",
    "VB":  "vb",
    "MÊS": "mes",
    "MES": "mes",
}


def normalise_unit(raw: str | None) -> str | None:
    if raw is None:
        return None
    key = raw.strip().upper()
    return UNIT_MAP.get(key, raw.strip().lower())



# ---------------------------------------------------------------------------
# Read input
# ---------------------------------------------------------------------------


@dataclass
class ColumnMapping:
    # Default indices (0-based): A=0, B=1, ...
    item_col: int = 1   # Col B
    desc_col: int = 2   # Col C
    code_col: int = 3   # Col D
    unit_col: int = 4   # Col E
    price_col: int = 5  # Col F
    qty_col: int = 18   # Col S
    start_row: int = 7

def read_input(file_obj, mapping: ColumnMapping = ColumnMapping(), sheet_name: str | None = None) -> list[InputItem]:
    """Read the input spreadsheet and return a list of InputItems."""
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook.")
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]

    # Indices are now directly provided in mapping (0-based)
    item_idx = mapping.item_col
    desc_idx = mapping.desc_col
    code_idx = mapping.code_col
    unit_idx = mapping.unit_col
    price_idx = mapping.price_col
    qty_idx = mapping.qty_col

    items: list[InputItem] = []
    
    # We iterate over all rows, but start at mapping.start_row
    for row in ws.iter_rows(min_row=mapping.start_row, max_row=ws.max_row, values_only=False):
        # row is a tuple of Cell objects
        # We need to ensure the row has enough columns for our indices
        max_idx = max(item_idx, desc_idx, code_idx, unit_idx, price_idx, qty_idx)
        if len(row) <= max_idx:
            continue

        item_cell = row[item_idx]
        desc_cell = row[desc_idx]
        code_cell = row[code_idx]
        unit_cell = row[unit_idx]
        price_cell = row[price_idx]
        qty_cell = row[qty_idx]

        item_val = item_cell.value
        if item_val is None:
            continue

        item_str = str(item_val).strip()
        if not item_str:
            continue

        desc = str(desc_cell.value or "").strip()
        code = code_cell.value
        unit = unit_cell.value
        price = price_cell.value
        qty = qty_cell.value

        # Determine if this is a data item (has code AND non-empty unit)
        has_code = code is not None and str(code).strip() != ""
        has_unit = unit is not None and str(unit).strip() not in ("", None)
        is_data = has_code and has_unit

        if code and isinstance(code, str):
            code = code.strip()
        if unit and isinstance(unit, str):
            unit = unit.strip()
            
        # Clean price and qty
        try:
             price_val = float(price) if is_data and price else None
        except (ValueError, TypeError):
             price_val = None
             
        try:
             qty_val = round(float(qty), 2) if is_data and qty else None
        except (ValueError, TypeError):
             qty_val = None

        items.append(InputItem(
            raw_item=item_str,
            description=desc,
            code=code if is_data else None,
            unit=unit if is_data else None,
            price=price_val,
            quantity=qty_val,
            is_data=is_data,
        ))

    wb.close()
    return items


# ---------------------------------------------------------------------------
# Transform
# ---------------------------------------------------------------------------

def transform(items: list[InputItem]) -> list[OutputRow]:
    """
    Transform items so every data item is at level 4.
    
    Rules:
    - Task identification: item.is_data (has code + unit).
    - Level 1: Always header.
    - Level 2 Task (e.g. 01.03):
        -> Create L3 header (01.03.001) using L2 desc.
        -> Create L4 task (01.03.001.001).
    - Level 3 Task (e.g. 10.03.04):
        -> Create L3 header (10.03.001) using L2 desc (if not exists).
        -> Create L4 task (10.03.001.04).
    - Level 4 Task: Keep as-is.
    - Level 5+ Task (e.g. 12.04.01.02.01):
        -> Flatten to Level 4 by merging parts from index 3 onwards.
        -> 12.04.01.02.01 -> 12.04.01.0201
    """
    output: list[OutputRow] = []

    # Map to store descriptions of potential parents
    # Key: "001", "001.002", etc.
    desc_map: dict[str, str] = {}
    
    # Track emitted synthetic headers to avoid duplicates
    emitted_headers: set[str] = set()

    for item in items:
        padded = item.padded_item
        parts = item.parts
        level = item.level
        
        # Always store description for lookup by children
        desc_map[padded] = item.description

        if level == 1:
            # Level 1 is always a group
            output.append(OutputRow(item=padded, description=item.description))
            
        elif level == 2:
            if not item.is_data:
                # L2 Group -> Keep
                output.append(OutputRow(item=padded, description=item.description))
            else:
                # L2 Task -> Needs to go to L4
                # Structure: L2 (Group) -> L3 (Group) -> L4 (Task)
                # But here L2 is the task itself. We need to create a wrapper structure?
                # Actually, if L2 is 01.03 and it's a task, it usually means 
                # "01.03 - My Task". 
                # Standard practice in this pipeline:
                # Create synthetic L3 header: 01.03.001 (desc = My Task)
                # Create L4 task: 01.03.001.001 (desc = My Task)
                
                # 1. Create synthetic L3 header
                l3_header = f"{padded}.001"
                if l3_header not in emitted_headers:
                    output.append(OutputRow(item=l3_header, description=item.description))
                    emitted_headers.add(l3_header)
                
                # 2. Create L4 task
                l4_item = f"{l3_header}.001"
                output.append(OutputRow(
                    item=l4_item,
                    code=item.code,
                    description=item.description,
                    unit=normalise_unit(item.unit),
                    quantity=item.quantity,
                    price=item.price,
                ))

        elif level == 3:
            if not item.is_data:
                # L3 Group -> Keep
                output.append(OutputRow(item=padded, description=item.description))
            else:
                # L3 Task (e.g. 10.03.04) -> Needs to go to L4
                # Parent L2 is 10.03
                parent_l2 = ".".join(parts[:2])
                
                # We need a Level 3 header to hold these L4 items.
                # Convention: Use parent.001 as the general container
                l3_header = f"{parent_l2}.001"
                
                if l3_header not in emitted_headers:
                    # Get description from L2 parent if possible
                    parent_desc = desc_map.get(parent_l2, item.description)
                    output.append(OutputRow(item=l3_header, description=parent_desc))
                    emitted_headers.add(l3_header)
                
                # Create L4 task: 10.03.001.04
                # Note: We use the original 3rd part (04) as the suffix
                suffix = parts[2]
                l4_item = f"{l3_header}.{suffix}"
                
                output.append(OutputRow(
                    item=l4_item,
                    code=item.code,
                    description=item.description,
                    unit=normalise_unit(item.unit),
                    quantity=item.quantity,
                    price=item.price,
                ))
                
        elif level == 4:
            # L4 (Task or Group) -> Output as is
            output.append(OutputRow(
                item=padded,
                code=item.code,
                description=item.description,
                unit=normalise_unit(item.unit),
                quantity=item.quantity,
                price=item.price,
            ))
            
        elif level >= 5:
            # Deep items -> Flatten to Level 4
            # Keep parts 0, 1, 2 (first 3 digits)
            # Merge parts 3 to end
            
            base = ".".join(parts[:3]) # 12.04.01
            
            # Merge remaining parts
            # e.g. parts[3] = "02", parts[4] = "01" -> "0201"
            rest = "".join(parts[3:])
            
            l4_item = f"{base}.{rest}"
            
            output.append(OutputRow(
                item=l4_item,
                code=item.code,
                description=item.description,
                unit=normalise_unit(item.unit),
                quantity=item.quantity,
                price=item.price,
            ))

    return output


# ---------------------------------------------------------------------------
# Write output
# ---------------------------------------------------------------------------

HEADER_ROW = [
    "ITEM",
    "CÓDIGO AUXILIAR",
    "DESCRIÇÃO DO SERVIÇO",
    "UNID.",
    "QUANTIDADE",
    "PREÇO UNITARIO",
]


def write_output(rows: list[OutputRow], file_obj_or_path) -> None:
    """Write the output spreadsheet to a file object or path."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Planilha"

    # Header
    ws.append(HEADER_ROW)

    for row in rows:
        ws.append([
            row.item,
            row.code,
            row.description,
            row.unit,
            row.quantity,
            row.price,
        ])

    wb.save(file_obj_or_path)
    wb.close()


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    if len(sys.argv) < 3:
        print("Usage: python pipeline.py <input.xlsx> <output.xlsx>")
        sys.exit(1)

    input_path = sys.argv[1]
    output_path = sys.argv[2]

    print(f"Reading: {input_path}")
    items = read_input(input_path)
    print(f"  {len(items)} items parsed")

    data_items = [it for it in items if it.is_data]
    header_items = [it for it in items if not it.is_data]
    print(f"  {len(data_items)} data items, {len(header_items)} headers")

    print("Transforming...")
    output_rows = transform(items)
    print(f"  {len(output_rows)} output rows")

    data_output = [r for r in output_rows if r.code is not None]
    print(f"  {len(data_output)} data rows in output")

    print(f"Writing: {output_path}")
    write_output(output_rows, output_path)
    print("Done!")


if __name__ == "__main__":
    main()
