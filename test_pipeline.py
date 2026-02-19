"""
Tests for pipeline.py

Covers:
  - normalise_unit (string, int, None inputs)
  - InputItem properties (parts, level, padded_item)
  - read_input (string and numeric cell values for code/unit)
  - transform (L2/L3/L4/L5+ task deepening / flattening)
  - Full round-trip: read → transform → write → re-read
"""

import pytest
import openpyxl
from io import BytesIO
from pipeline import (
    InputItem,
    OutputRow,
    ColumnMapping,
    normalise_unit,
    read_input,
    transform,
    write_output,
)


# ═══════════════════════════════════════════════════════════════════════════
# Helpers
# ═══════════════════════════════════════════════════════════════════════════

def _make_workbook(rows, start_row=1):
    """Create an in-memory Excel workbook with the given rows.

    Each row is a list of cell values. Rows are written starting at `start_row`.
    Returns a BytesIO buffer ready for read_input().
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    for r_idx, row in enumerate(rows, start=start_row):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    buf = BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)
    return buf


def _default_mapping(**overrides):
    """ColumnMapping with 0-based indices matching our test workbook layout.

    Layout: col0=ITEM, col1=DESC, col2=CODE, col3=UNIT, col4=PRICE, col5=QTY
    """
    defaults = dict(
        item_col=0,
        desc_col=1,
        code_col=2,
        unit_col=3,
        price_col=4,
        qty_col=5,
        start_row=1,
    )
    defaults.update(overrides)
    return ColumnMapping(**defaults)


# ═══════════════════════════════════════════════════════════════════════════
# normalise_unit
# ═══════════════════════════════════════════════════════════════════════════

class TestNormaliseUnit:
    def test_none_returns_none(self):
        assert normalise_unit(None) is None

    def test_known_units(self):
        assert normalise_unit("M2") == "m2"
        assert normalise_unit("m2") == "m2"
        assert normalise_unit("UND") == "un"
        assert normalise_unit("VB") == "vb"
        assert normalise_unit("MÊS") == "mes"

    def test_unknown_unit_lowercased(self):
        assert normalise_unit("CX") == "cx"
        assert normalise_unit("Pç") == "pç"

    def test_whitespace_stripped(self):
        assert normalise_unit("  M2  ") == "m2"
        assert normalise_unit("  VB") == "vb"

    def test_integer_input_does_not_crash(self):
        """Regression: .strip() on int should not raise."""
        result = normalise_unit(1)
        assert isinstance(result, str)
        assert result == "1"

    def test_float_input_does_not_crash(self):
        result = normalise_unit(2.5)
        assert isinstance(result, str)


# ═══════════════════════════════════════════════════════════════════════════
# InputItem
# ═══════════════════════════════════════════════════════════════════════════

class TestInputItem:
    def test_single_part(self):
        item = InputItem(raw_item="1", description="Obras", code=None, unit=None, price=None, quantity=None)
        assert item.parts == ["001"]
        assert item.level == 1
        assert item.padded_item == "001"

    def test_two_parts(self):
        item = InputItem(raw_item="2.03", description="Sub", code=None, unit=None, price=None, quantity=None)
        assert item.parts == ["002", "003"]
        assert item.level == 2

    def test_four_parts_padded(self):
        item = InputItem(raw_item="1.2.3.4", description="T", code="C", unit="M2", price=10.0, quantity=1.0, is_data=True)
        assert item.padded_item == "001.002.003.004"
        assert item.level == 4

    def test_five_parts(self):
        item = InputItem(raw_item="12.04.01.02.01", description="D", code="C", unit="M2", price=1.0, quantity=1.0, is_data=True)
        assert item.level == 5
        assert item.padded_item == "012.004.001.002.001"


# ═══════════════════════════════════════════════════════════════════════════
# read_input – type handling (the .strip() bug)
# ═══════════════════════════════════════════════════════════════════════════

class TestReadInput:
    def test_string_code_and_unit(self):
        """Standard case: code and unit are strings."""
        buf = _make_workbook([
            # ITEM, DESC,     CODE,    UNIT,  PRICE, QTY
            ["1",   "Header", None,    None,  None,  None],
            ["1.1", "Task A", "ABC01", "M2",  10.0,  5.0],
        ])
        mapping = _default_mapping()
        items = read_input(buf, mapping)

        assert len(items) == 2
        header, task = items
        assert not header.is_data
        assert task.is_data
        assert task.code == "ABC01"
        assert task.unit == "M2"
        assert task.price == 10.0
        assert task.quantity == 5.0

    def test_numeric_code_and_unit(self):
        """Regression: code and unit are integers in Excel → should not crash."""
        buf = _make_workbook([
            ["1.1.1.1", "Task B", 12345, 2, 99.0, 3.0],
        ])
        mapping = _default_mapping()
        items = read_input(buf, mapping)

        assert len(items) == 1
        task = items[0]
        assert task.is_data is True
        assert task.code == "12345"
        assert task.unit == "2"

    def test_float_code_and_unit(self):
        """Regression: code/unit are floats in Excel."""
        buf = _make_workbook([
            ["1.1.1.1", "Task C", 3.14, 1.5, 50.0, 2.0],
        ])
        mapping = _default_mapping()
        items = read_input(buf, mapping)

        assert len(items) == 1
        task = items[0]
        assert task.is_data is True
        assert isinstance(task.code, str)
        assert isinstance(task.unit, str)

    def test_empty_code_is_not_data(self):
        """Rows with empty code should not be marked as data items."""
        # Note: trailing None cells may be omitted by openpyxl, so we put 0
        # in the last column to ensure rows have enough columns.
        buf = _make_workbook([
            ["1", "Header", None, None, None, 0],
            ["1.1", "Sub Header", "", "M2", None, 0],
        ])
        mapping = _default_mapping()
        items = read_input(buf, mapping)

        assert len(items) == 2
        assert not items[0].is_data
        assert not items[1].is_data  # empty code → not data

    def test_skips_blank_item_rows(self):
        """Rows with None in the ITEM column should be skipped."""
        buf = _make_workbook([
            ["1", "Header", None, None, None, None],
            [None, "Ignored", "X", "M2", 1.0, 1.0],
            ["1.1", "Task", "C1", "M2", 10.0, 5.0],
        ])
        mapping = _default_mapping()
        items = read_input(buf, mapping)
        assert len(items) == 2  # skipped the None row

    def test_sheet_selection(self):
        """read_input should use the selected sheet."""
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Ignored"
        ws1.cell(1, 1, "1")
        ws1.cell(1, 2, "Wrong Sheet")

        ws2 = wb.create_sheet("Target")
        ws2.cell(1, 1, "1")
        ws2.cell(1, 2, "Right Sheet")
        ws2.cell(1, 3, "C1")
        ws2.cell(1, 4, "M2")
        ws2.cell(1, 5, 10.0)
        ws2.cell(1, 6, 1.0)

        buf = BytesIO()
        wb.save(buf)
        wb.close()
        buf.seek(0)

        mapping = _default_mapping()
        items = read_input(buf, mapping, sheet_name="Target")
        assert len(items) == 1
        assert items[0].description == "Right Sheet"


# ═══════════════════════════════════════════════════════════════════════════
# transform
# ═══════════════════════════════════════════════════════════════════════════

class TestTransform:
    def _make_items(self, specs):
        """Helper: build InputItem list from (raw_item, desc, is_data, code, unit, price, qty) tuples."""
        items = []
        for raw, desc, data, code, unit, price, qty in specs:
            items.append(InputItem(
                raw_item=raw, description=desc,
                code=code, unit=unit, price=price, quantity=qty, is_data=data,
            ))
        return items

    def test_l4_task_kept_as_is(self):
        """A level-4 task should stay at level 4."""
        items = self._make_items([
            ("1",       "Nível 1",   False, None, None, None, None),
            ("1.1",     "Nível 2",   False, None, None, None, None),
            ("1.1.1",   "Nível 3",   False, None, None, None, None),
            ("1.1.1.1", "Tarefa L4", True,  "C1", "M2", 10.0, 5.0),
        ])
        rows = transform(items)
        data_rows = [r for r in rows if r.code is not None]
        assert len(data_rows) == 1
        assert data_rows[0].item == "001.001.001.001"
        assert data_rows[0].unit == "m2"

    def test_l2_task_deepened_to_l4(self):
        """A level-2 task should get synthetic L2 + L3 headers and move to L4."""
        items = self._make_items([
            ("1",   "Parent L1", False, None, None, None, None),
            ("1.1", "Task at L2", True,  "C1", "UND", 20.0, 3.0),
        ])
        rows = transform(items)
        # Should have: L1 header, syn L2, syn L3, L4 task
        assert len(rows) == 4
        assert rows[0].item == "001"           # L1 header
        assert rows[1].item == "001.001"       # synthetic L2
        assert rows[2].item == "001.001.001"   # synthetic L3
        assert rows[3].item == "001.001.001.001"  # actual task
        assert rows[3].code == "C1"

    def test_l3_task_deepened_to_l4(self):
        """A level-3 task should get a synthetic L3 header and move to L4."""
        items = self._make_items([
            ("1",     "Parent L1",  False, None, None, None, None),
            ("1.1",   "Parent L2",  False, None, None, None, None),
            ("1.1.1", "Task at L3", True,  "C2", "KG", 5.0, 100.0),
        ])
        rows = transform(items)
        # L1, L2 headers, syn L3, L4 task = 4 rows
        data_rows = [r for r in rows if r.code is not None]
        assert len(data_rows) == 1
        assert data_rows[0].item.count(".") == 3  # 4 levels

    def test_l5_task_flattened(self):
        """A level-5+ task should be flattened to level 4 with merged suffix."""
        items = self._make_items([
            ("1",           "L1",    False, None, None, None, None),
            ("1.1",         "L2",    False, None, None, None, None),
            ("1.1.1",       "L3",    False, None, None, None, None),
            ("1.1.1.1.1",   "Deep",  True,  "C3", "M3", 15.0, 2.0),
        ])
        rows = transform(items)
        data_rows = [r for r in rows if r.code is not None]
        assert len(data_rows) == 1
        # L5 item 001.001.001.001.001 → flattened to 001.001.001.001001
        assert data_rows[0].item.startswith("001.001.001.")
        parts = data_rows[0].item.split(".")
        assert len(parts) == 4  # still 4 levels

    def test_sequential_numbering_under_l3(self):
        """Multiple tasks under the same L3 container get sequential numbers."""
        items = self._make_items([
            ("1",       "L1",     False, None, None, None, None),
            ("1.1",     "L2",     False, None, None, None, None),
            ("1.1.1",   "L3",     False, None, None, None, None),
            ("1.1.1.1", "Task 1", True,  "C1", "M2", 10.0, 1.0),
            ("1.1.1.2", "Task 2", True,  "C2", "M2", 20.0, 2.0),
            ("1.1.1.3", "Task 3", True,  "C3", "M2", 30.0, 3.0),
        ])
        rows = transform(items)
        data_rows = [r for r in rows if r.code is not None]
        assert len(data_rows) == 3
        assert data_rows[0].item == "001.001.001.001"
        assert data_rows[1].item == "001.001.001.002"
        assert data_rows[2].item == "001.001.001.003"

    def test_numeric_unit_transform_no_crash(self):
        """Regression: numeric unit values flow through transform without error."""
        items = self._make_items([
            ("1",       "L1",    False, None, None, None, None),
            ("1.1",     "L2",    False, None, None, None, None),
            ("1.1.1",   "L3",    False, None, None, None, None),
            ("1.1.1.1", "Task",  True,  "123", "2", 10.0, 5.0),
        ])
        # Should not raise
        rows = transform(items)
        data_rows = [r for r in rows if r.code is not None]
        assert len(data_rows) == 1
        assert data_rows[0].unit == "2"  # normalise_unit passes through unknown units


# ═══════════════════════════════════════════════════════════════════════════
# Round-trip: read → transform → write → re-read
# ═══════════════════════════════════════════════════════════════════════════

class TestRoundTrip:
    def test_write_and_reopen(self):
        """write_output produces a valid .xlsx that can be reopened."""
        rows = [
            OutputRow(item="001", description="Header"),
            OutputRow(item="001.001", description="Sub"),
            OutputRow(item="001.001.001", description="Sub-sub"),
            OutputRow(item="001.001.001.001", code="C1", description="Task", unit="m2", quantity=5.0, price=10.0),
        ]
        buf = BytesIO()
        write_output(rows, buf)
        buf.seek(0)

        wb = openpyxl.load_workbook(buf)
        ws = wb.active
        # Header + 4 data rows = 5 rows
        all_rows = list(ws.iter_rows(values_only=True))
        assert len(all_rows) == 5
        assert all_rows[0][0] == "ITEM"  # header
        assert all_rows[4][1] == "C1"    # code in last row
        wb.close()

    def test_full_pipeline_with_numeric_values(self):
        """End-to-end: Excel with numeric code/unit → transform → output, no crash."""
        buf = _make_workbook([
            # ITEM,       DESC,        CODE,  UNIT,  PRICE, QTY
            ["1",         "Obras",      None,  None,  None,  None],
            ["1.1",       "Estrutura",  None,  None,  None,  None],
            ["1.1.1",     "Concreto",   None,  None,  None,  None],
            ["1.1.1.1",   "Tarefa A",   101,   2,     50.0,  10.0],   # numeric code + unit
            ["1.1.1.2",   "Tarefa B",   "XY",  "M2",  30.0,  5.0],   # string code + unit
        ])
        mapping = _default_mapping()
        items = read_input(buf, mapping)
        assert len(items) == 5

        output = transform(items)
        data_rows = [r for r in output if r.code is not None]
        assert len(data_rows) == 2
        # Both should have string codes
        assert all(isinstance(r.code, str) for r in data_rows)


# ═══════════════════════════════════════════════════════════════════════════
# Integration with real workbooks (if available)
# ═══════════════════════════════════════════════════════════════════════════

class TestIntegrationRealFiles:
    """Run against real input files if they exist. Skipped otherwise."""

    @pytest.fixture(autouse=True)
    def _check_files(self):
        from pathlib import Path
        self.project = Path(__file__).parent
        self.camil = self.project / "planilha_camil.xlsx"
        self.leonardo = self.project / "planilha_leonardo.xlsx"

    def test_camil_no_crash(self):
        if not self.camil.exists():
            pytest.skip("planilha_camil.xlsx not present")
        mapping = ColumnMapping(item_col=1, desc_col=3, code_col=4, unit_col=5, price_col=6, qty_col=7, start_row=7)
        items = read_input(str(self.camil), mapping)
        assert len(items) > 0
        output = transform(items)
        assert len(output) > 0

    def test_leonardo_no_crash(self):
        if not self.leonardo.exists():
            pytest.skip("planilha_leonardo.xlsx not present")
        mapping = ColumnMapping()
        items = read_input(str(self.leonardo), mapping)
        assert len(items) > 0
        output = transform(items)
        assert len(output) > 0
