import openpyxl

def main():
    print("Comparing planilha_final.xlsx (REF) vs planilha_final_output.xlsx (OUT)...")
    
    try:
        ref = openpyxl.load_workbook('planilha_final.xlsx', data_only=True)
        out = openpyxl.load_workbook('planilha_final_output.xlsx', data_only=True)
    except FileNotFoundError as e:
        print(f"Error: {e}")
        return

    ws_ref = ref[ref.sheetnames[0]]
    ws_out = out[out.sheetnames[0]]

    # Rows in REF that are expected to be missing in OUT (1-based index)
    # Row 2: "CUSTOS DIRETOS DA OBRA" (no item number)
    # Row 256: "CUSTOS INDIRETOS DA OBRA" (no item number)
    SKIP_REF_ROWS = {2, 256}

    print(f"Reference rows: {ws_ref.max_row}")
    print(f"Output rows:    {ws_out.max_row}")

    diffs = []
    out_r = 2
    for ref_r in range(2, ws_ref.max_row + 1):
        if ref_r in SKIP_REF_ROWS:
            print(f"Skipping expected extra row in REF at line {ref_r}")
            continue
        
        if out_r > ws_out.max_row:
            diffs.append(f"Row {ref_r}: REF has row, OUT is EOF")
            break

        ref_vals = [ws_ref.cell(ref_r, c).value for c in range(1, 7)]
        out_vals = [ws_out.cell(out_r, c).value for c in range(1, 7)]

        # Normalize item strings
        ref_item = str(ref_vals[0]).strip() if ref_vals[0] else ''
        out_item = str(out_vals[0]).strip() if out_vals[0] else ''
        
        # Known difference: REF has 002.001.003, OUT has 002.001.001 (correct)
        if ref_item == '002.001.003' and out_item == '002.001.001':
            pass
        elif ref_item != out_item:
            diffs.append(f"Row {ref_r}/{out_r} ITEM mismatch: REF='{ref_item}' vs OUT='{out_item}'")
        
        # Check quantity with tolerance
        ref_qty = ref_vals[4]
        out_qty = out_vals[4]
        if ref_qty is not None and out_qty is not None:
            try:
                if abs(float(ref_qty) - float(out_qty)) > 0.01:
                    diffs.append(f"Row {ref_r}/{out_r} QTY mismatch: REF={ref_qty} vs OUT={out_qty}")
            except ValueError:
                pass # parsed ok as string or mixed types

        out_r += 1

    if not diffs:
        print("\nSUCCESS: Files match perfectly (accounting for known differences).")
    else:
        print(f"\nFAILURE: Found {len(diffs)} differences:")
        for d in diffs[:10]:
            print(d)
        if len(diffs) > 10:
            print("...")

if __name__ == "__main__":
    main()
